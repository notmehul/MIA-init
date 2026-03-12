from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


TAB_BUILD_ORDER = [
    ("context/deal-context.json", "Pre-Requisites"),
    ("sizing/top-down.json", "Top Down"),
    ("sizing/bottoms-up.json", "Bottoms Up"),
    ("analysis/value-chain.json", "MI: Value Chain"),
    ("analysis/trends.json", "MI: Trends"),
    ("analysis/competitors.json", "Competitor Analysis"),
    ("analysis/feature-comparison.json", "Feature Comparison"),
    ("analysis/meta-review.json", "MI: Meta Review"),
    ("validation/whale-watch.json", "MI: Whale Watch"),
    ("validation/anti-feku.json", "MI: Anti-Feku"),
]

ALLOWED_ROW_TYPES = {"data", "section_header", "total", "empty"}
INVALID_SHEET_TITLE_CHARS = re.compile(r"[\\/*?:\[\]]")

header_font = Font(bold=True, size=11, name="Arial", color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E79")
section_fill = PatternFill("solid", fgColor="D6E4F0")
section_font = Font(bold=True, size=11, name="Arial", color="1F4E79")
data_font = Font(size=10, name="Arial")
bold_font = Font(bold=True, size=10, name="Arial")
warn_font = Font(size=10, name="Arial", color="CC6600")
summary_font = Font(size=10, name="Arial", italic=True)
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
wrap = Alignment(wrap_text=True, vertical="top")
header_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _load_json(path: Path) -> Any:
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def _require_mapping(value: Any, label: str, path: Path) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise ValueError(f"{path}: expected {label} to be an object")
    return value


def _require_list(value: Any, label: str, path: Path) -> list[Any]:
    if not isinstance(value, list):
        raise ValueError(f"{path}: expected {label} to be a list")
    return value


def load_workspace_meta(workspace_dir: Path) -> dict[str, Any]:
    meta_path = workspace_dir / "meta.json"
    meta = _require_mapping(_load_json(meta_path), "meta.json", meta_path)
    for key in ("deal_slug", "company_name", "workspace_version"):
        if not meta.get(key):
            raise ValueError(f"{meta_path}: missing required field '{key}'")
    return meta


def load_sources(workspace_dir: Path) -> dict[str, Any]:
    sources_path = workspace_dir / "sources.json"
    if not sources_path.exists():
        return {"sources": {}}

    payload = _require_mapping(_load_json(sources_path), "sources.json", sources_path)
    sources = payload.get("sources", {})
    if not isinstance(sources, dict):
        raise ValueError(f"{sources_path}: expected 'sources' to be an object")

    for source_id, source in sources.items():
        _require_mapping(source, f"source '{source_id}'", sources_path)

    return payload


def validate_skill_payload(payload: dict[str, Any], payload_path: Path, sources: dict[str, Any]) -> None:
    for key in ("skill", "tab_name", "summary", "tables"):
        if key not in payload:
            raise ValueError(f"{payload_path}: missing required field '{key}'")

    tables = _require_list(payload["tables"], "tables", payload_path)
    if not tables:
        raise ValueError(f"{payload_path}: expected at least one table")

    source_ids = set(sources.get("sources", {}).keys())
    for table in tables:
        validate_table(table, payload_path, source_ids)

    sections = payload.get("sections", [])
    if sections is not None:
        _require_list(sections, "sections", payload_path)

    top_level_refs = payload.get("source_refs", [])
    if top_level_refs is not None:
        _validate_source_refs(top_level_refs, source_ids, payload_path, "top-level source_refs")


def validate_table(table: Any, payload_path: Path, source_ids: set[str]) -> None:
    table_data = _require_mapping(table, "table", payload_path)
    for key in ("id", "label", "headers", "rows"):
        if key not in table_data:
            raise ValueError(f"{payload_path}: table missing required field '{key}'")

    headers = _require_list(table_data["headers"], "table headers", payload_path)
    if not headers:
        raise ValueError(f"{payload_path}: table '{table_data['id']}' must have at least one header")

    rows = _require_list(table_data["rows"], "table rows", payload_path)
    for row_index, row in enumerate(rows):
        validate_row(row, len(headers), payload_path, str(table_data["id"]), row_index, source_ids)


def validate_row(
    row: Any,
    column_count: int,
    payload_path: Path,
    table_id: str,
    row_index: int,
    source_ids: set[str],
) -> None:
    row_data = _require_mapping(row, f"row {row_index}", payload_path)
    row_type = row_data.get("type", "data")
    if row_type not in ALLOWED_ROW_TYPES:
        raise ValueError(
            f"{payload_path}: table '{table_id}' row {row_index} has unsupported type '{row_type}'"
        )

    data = row_data.get("data", [])
    if row_type in {"data", "total"}:
        if not isinstance(data, list) or len(data) != column_count:
            raise ValueError(
                f"{payload_path}: table '{table_id}' row {row_index} must have {column_count} cells"
            )
    elif row_type == "section_header":
        has_data_text = isinstance(data, list) and any(str(cell).strip() for cell in data if cell is not None)
        if not has_data_text:
            raise ValueError(
                f"{payload_path}: table '{table_id}' row {row_index} section_header needs a non-empty first cell in data"
            )

    refs = row_data.get("source_refs", [])
    _validate_source_refs(refs, source_ids, payload_path, f"table '{table_id}' row {row_index}")


def _validate_source_refs(refs: Any, source_ids: set[str], payload_path: Path, label: str) -> None:
    if refs is None:
        return
    ref_list = _require_list(refs, label, payload_path)
    for ref in ref_list:
        if not isinstance(ref, str):
            raise ValueError(f"{payload_path}: {label} entries must be strings")
        if source_ids and ref not in source_ids:
            raise ValueError(f"{payload_path}: {label} references unknown source '{ref}'")


def make_sheet_title(raw_title: str, used_titles: set[str]) -> str:
    sanitized = INVALID_SHEET_TITLE_CHARS.sub(" -", raw_title).strip() or "Sheet"
    base = sanitized[:31]
    candidate = base
    suffix = 2
    while candidate in used_titles:
        tail = f" ({suffix})"
        candidate = f"{base[: 31 - len(tail)]}{tail}"
        suffix += 1
    return candidate


def sanitize_filename(company_name: str) -> str:
    return company_name.replace("/", "-").replace("\\", "-").strip() or "Market_Intelligence"


class TabBuilder:
    def __init__(self, worksheet, sources: dict[str, Any]) -> None:
        self.ws = worksheet
        self.sources = sources.get("sources", {})
        self.row = 1
        self.max_columns = 1

    def build(self, payload: dict[str, Any]) -> None:
        tables = payload["tables"]
        self.max_columns = max(len(table["headers"]) for table in tables)

        self.write_summary(payload["summary"], self.max_columns)
        self.row += 1

        for table in tables:
            self.write_table(table)
            self.row += 1

        for section in payload.get("sections", []) or []:
            self.write_section_block(section)
            self.row += 1

        self.ws.freeze_panes = "A2"

    def write_table(self, table: dict[str, Any]) -> None:
        headers = table["headers"]
        num_cols = len(headers)
        label = table.get("label")
        if label:
            self.write_section_header(label, num_cols)
            self.row += 1

        self.write_header_row(headers, num_cols)
        self.row += 1

        source_col = self._find_source_column(headers)
        notes_col = self._find_agent_notes_column(headers)
        for row_payload in table["rows"]:
            row_type = row_payload.get("type", "data")
            if row_type == "section_header":
                text = next(
                    (str(cell) for cell in row_payload.get("data", []) if cell not in (None, "")),
                    "",
                )
                self.write_section_header(text, num_cols)
                self.row += 1
                continue

            if row_type == "empty":
                self.row += 1
                continue

            is_bold = row_type == "total"
            self.write_data_row(row_payload.get("data", []), num_cols, is_bold, notes_col)
            self.apply_source_hyperlinks(source_col, row_payload.get("source_refs", []))
            self.row += 1

    def write_summary(self, text: str, num_cols: int) -> None:
        self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=num_cols)
        cell = self.ws.cell(row=self.row, column=1)
        cell.value = text
        cell.font = summary_font
        cell.alignment = wrap

    def write_header_row(self, headers: list[Any], num_cols: int) -> None:
        for column, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=self.row, column=column)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for column in range(len(headers) + 1, num_cols + 1):
            cell = self.ws.cell(row=self.row, column=column)
            cell.border = thin_border

    def write_data_row(self, row_data: list[Any], num_cols: int, is_bold: bool, notes_col: int | None) -> None:
        for column, value in enumerate(row_data, start=1):
            cell = self.ws.cell(row=self.row, column=column)
            cell.value = value
            cell.font = self._cell_font(value, is_bold, notes_col, column)
            cell.alignment = wrap
            cell.border = thin_border

        for column in range(len(row_data) + 1, num_cols + 1):
            cell = self.ws.cell(row=self.row, column=column)
            cell.font = bold_font if is_bold else data_font
            cell.alignment = wrap
            cell.border = thin_border

    def write_section_header(self, text: str, num_cols: int) -> None:
        self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=num_cols)
        cell = self.ws.cell(row=self.row, column=1)
        cell.value = text
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = wrap
        for column in range(1, num_cols + 1):
            self.ws.cell(row=self.row, column=column).border = thin_border

    def write_section_block(self, section: dict[str, Any]) -> None:
        num_cols = self.max_columns
        label = section.get("label")
        if label:
            self.write_section_header(label, num_cols)
            self.row += 1

        text = section.get("content") or section.get("text")
        if text:
            self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=num_cols)
            cell = self.ws.cell(row=self.row, column=1)
            cell.value = text
            cell.font = data_font
            cell.alignment = wrap
            self.row += 1

        items = section.get("items", [])
        if isinstance(items, list):
            for item in items:
                self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=num_cols)
                cell = self.ws.cell(row=self.row, column=1)
                cell.value = f"- {item}"
                cell.font = data_font
                cell.alignment = wrap
                self.row += 1

        self.row -= 1

    def apply_source_hyperlinks(self, source_col: int | None, source_refs: list[str]) -> None:
        if source_col is None or not source_refs:
            return

        urls = []
        for source_ref in source_refs:
            source = self.sources.get(source_ref, {})
            url = source.get("url")
            if url:
                urls.append(url)

        if not urls:
            return

        cell = self.ws.cell(row=self.row, column=source_col)
        cell.hyperlink = urls[0]
        cell.style = "Hyperlink"

        extra_urls = urls[1:]
        if extra_urls:
            existing_text = f"{cell.comment.text}\n\n" if cell.comment else ""
            comment_text = f"{existing_text}Additional source URLs:\n" + "\n".join(extra_urls)
            cell.comment = Comment(comment_text, "Biome MI")

    def _find_source_column(self, headers: list[Any]) -> int | None:
        for index, header in enumerate(headers, start=1):
            if "source" in str(header).lower():
                return index
        return None

    def _find_agent_notes_column(self, headers: list[Any]) -> int | None:
        for index, header in enumerate(headers, start=1):
            if str(header).strip().lower() == "agent notes":
                return index
        return None

    def _cell_font(self, value: Any, is_bold: bool, notes_col: int | None, column: int) -> Font:
        if notes_col == column and value and "⚠" in str(value):
            return warn_font
        return bold_font if is_bold else data_font


def build_workbook(workspace_dir: str | Path) -> Path:
    workspace_path = Path(workspace_dir).resolve()
    meta = load_workspace_meta(workspace_path)
    sources = load_sources(workspace_path)

    workbook = Workbook()
    default_sheet = workbook.active
    used_titles: set[str] = set()
    sheets_built = 0

    for relative_path, default_title in TAB_BUILD_ORDER:
        payload_path = workspace_path / relative_path
        if not payload_path.exists():
            continue

        payload = _require_mapping(_load_json(payload_path), str(payload_path), payload_path)
        validate_skill_payload(payload, payload_path, sources)

        raw_title = str(payload.get("tab_name") or default_title)
        sheet_title = make_sheet_title(raw_title, used_titles)
        worksheet = default_sheet if sheets_built == 0 else workbook.create_sheet()
        worksheet.title = sheet_title
        used_titles.add(sheet_title)

        TabBuilder(worksheet, sources).build(payload)
        sheets_built += 1

    if sheets_built == 0:
        raise ValueError(f"{workspace_path}: no skill JSON files found in the expected MI v3 locations")

    output_dir = workspace_path / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_name = f"{sanitize_filename(str(meta['company_name']))}_Market_Intelligence.xlsx"
    output_path = output_dir / output_name
    workbook.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a Biome MI v3 workbook from a workspace directory.")
    parser.add_argument("workspace_dir", help="Path to the MI v3 workspace directory")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_path = build_workbook(args.workspace_dir)
    print(output_path)


if __name__ == "__main__":
    main()
