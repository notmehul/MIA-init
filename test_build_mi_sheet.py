from __future__ import annotations

import importlib.util
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).with_name("build_mi_sheet.py")
SPEC = importlib.util.spec_from_file_location("build_mi_sheet", SCRIPT_PATH)
MODULE = importlib.util.module_from_spec(SPEC)
assert SPEC and SPEC.loader
SPEC.loader.exec_module(MODULE)


TAB_FILES = [
    ("context/deal-context.json", "Pre-Requisites"),
    ("sizing/top-down.json", "Top Down"),
    ("sizing/bottoms-up.json", "Bottoms Up"),
    ("analysis/value-chain.json", "MI: Value Chain"),
    ("analysis/trends.json", "MI: Trends"),
    ("analysis/competitors.json", "Competitor Analysis"),
    ("analysis/feature-comparison.json", "Feature Comparison"),
    ("analysis/meta-review.json", "MI: Meta Review"),
    ("validation/whale-watch.json", "MI: Whale Watch"),
    ("validation/anti-feku.json", "MI: Anti-Feku"),
]


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def make_workspace() -> Path:
    workspace = Path(tempfile.mkdtemp(prefix="mi-v3-workspace-"))
    write_json(
        workspace / "meta.json",
        {
            "deal_slug": "test-deal",
            "company_name": "Test Company",
            "workspace_version": "3.0",
        },
    )
    write_json(
        workspace / "sources.json",
        {
            "deal_slug": "test-deal",
            "sources": {
                "src_001": {
                    "url": "https://example.com/source-1",
                    "publisher": "Example Research",
                    "title": "Source One",
                    "tier": 2,
                    "date_accessed": "2026-03-10",
                    "used_by_skills": [],
                },
                "src_002": {
                    "url": "https://example.com/source-2",
                    "publisher": "Example Research",
                    "title": "Source Two",
                    "tier": 2,
                    "date_accessed": "2026-03-10",
                    "used_by_skills": [],
                },
            },
        },
    )
    return workspace


def make_payload(tab_name: str, skill_name: str) -> dict:
    return {
        "skill": skill_name,
        "tab_name": tab_name,
        "summary": f"Summary for {skill_name}",
        "tables": [
            {
                "id": "table-1",
                "label": "Primary Table",
                "headers": ["Name", "Source", "Agent Notes"],
                "rows": [
                    {
                        "type": "data",
                        "data": ["Row 1", "(Source) - Example Research", ""],
                        "source_refs": ["src_001", "src_002"],
                    },
                    {
                        "type": "section_header",
                        "data": ["Section Divider", "", ""],
                    },
                    {
                        "type": "total",
                        "data": ["Total", "Calculated field", ""],
                        "source_refs": [],
                    },
                    {
                        "type": "data",
                        "data": ["Warning Row", "(Source) - Example Research", "⚠ Needs review"],
                        "source_refs": ["src_001"],
                    },
                ],
            }
        ],
        "sections": [
            {
                "id": "notes",
                "label": "Gut Check",
                "content": "This is a short conclusion.",
            }
        ],
        "source_refs": ["src_001"],
    }


def find_row(worksheet, first_cell_value: str) -> int:
    for row in worksheet.iter_rows():
        if row[0].value == first_cell_value:
            return row[0].row
    raise AssertionError(f"Could not find row starting with '{first_cell_value}'")


class BuildMiSheetTests(unittest.TestCase):
    def test_builds_full_workspace_in_expected_order(self) -> None:
        workspace = make_workspace()
        for relative_path, tab_name in TAB_FILES:
            skill_name = Path(relative_path).stem
            write_json(workspace / relative_path, make_payload(tab_name, skill_name))

        output_path = MODULE.build_workbook(workspace)
        workbook = load_workbook(output_path)

        self.assertEqual(
            workbook.sheetnames,
            [
                "Pre-Requisites",
                "Top Down",
                "Bottoms Up",
                "MI - Value Chain",
                "MI - Trends",
                "Competitor Analysis",
                "Feature Comparison",
                "MI - Meta Review",
                "MI - Whale Watch",
                "MI - Anti-Feku",
            ],
        )

    def test_skips_missing_files_for_partial_workspace(self) -> None:
        workspace = make_workspace()
        for relative_path, tab_name in TAB_FILES[:3]:
            skill_name = Path(relative_path).stem
            write_json(workspace / relative_path, make_payload(tab_name, skill_name))

        output_path = MODULE.build_workbook(workspace)
        workbook = load_workbook(output_path)

        self.assertEqual(workbook.sheetnames, ["Pre-Requisites", "Top Down", "Bottoms Up"])

    def test_applies_hyperlinks_comments_and_styles(self) -> None:
        workspace = make_workspace()
        write_json(workspace / "sizing/top-down.json", make_payload("Top Down", "market-sizing-top-down"))

        output_path = MODULE.build_workbook(workspace)
        worksheet = load_workbook(output_path)["Top Down"]

        row_1 = find_row(worksheet, "Row 1")
        source_cell = worksheet.cell(row=row_1, column=2)
        self.assertEqual(source_cell.hyperlink.target, "https://example.com/source-1")
        self.assertIn("https://example.com/source-2", source_cell.comment.text)

        warning_row = find_row(worksheet, "Warning Row")
        warning_cell = worksheet.cell(row=warning_row, column=3)
        self.assertIn("CC6600", warning_cell.font.color.rgb)

        total_row = find_row(worksheet, "Total")
        self.assertTrue(worksheet.cell(row=total_row, column=1).font.bold)

        merged_ranges = {str(cell_range) for cell_range in worksheet.merged_cells.ranges}
        self.assertIn("A5:C5", merged_ranges)

    def test_handles_empty_sources_registry(self) -> None:
        workspace = make_workspace()
        write_json(
            workspace / "sources.json",
            {
                "deal_slug": "test-deal",
                "sources": {},
            },
        )
        payload = make_payload("Top Down", "market-sizing-top-down")
        payload["tables"][0]["rows"][0]["source_refs"] = []
        payload["tables"][0]["rows"][3]["source_refs"] = []
        payload["source_refs"] = []
        write_json(workspace / "sizing/top-down.json", payload)

        output_path = MODULE.build_workbook(workspace)
        worksheet = load_workbook(output_path)["Top Down"]
        row_1 = find_row(worksheet, "Row 1")
        self.assertIsNone(worksheet.cell(row=row_1, column=2).hyperlink)

    def test_rejects_malformed_rows(self) -> None:
        workspace = make_workspace()
        payload = make_payload("Top Down", "market-sizing-top-down")
        payload["tables"][0]["rows"][0]["data"] = ["too", "short"]
        write_json(workspace / "sizing/top-down.json", payload)

        with self.assertRaisesRegex(ValueError, "must have 3 cells"):
            MODULE.build_workbook(workspace)


if __name__ == "__main__":
    unittest.main()
